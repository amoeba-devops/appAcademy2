#!/usr/bin/env python3
"""
Generate seed SQL for amb_acm_dsh_daily_kpi from
docs/reference/[TPI] Master.xlsx -> sheet[0] (INDEX).

Run:
  python3 scripts/build-dsh-index-seed.py

Output:
  sql/acm/700-seed-dsh-index-data.sql

Year baseline: row5 has Day=1, MS='목' for the first listed month '12월'.
2022-12-01 is Thursday -> baseline year = 2022.
"""
from __future__ import annotations

import datetime as dt
import os
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs/reference/[TPI] Master.xlsx"
OUT = ROOT / "sql/acm/700-seed-dsh-index-data.sql"
ENT_ID = "00000000-0000-0000-0000-000000000001"

DOW_KR_TO_EN = {"일": "SUN", "월": "MON", "화": "TUE", "수": "WED", "목": "THU", "금": "FRI", "토": "SAT"}

# Month start row map -> (sheet row of first day of month, month_label, year, month)
MONTHS = [
    (5,   "12월", 2022, 12),
    (38,  "1월",  2023, 1),
    (68,  "2월",  2023, 2),
    (101, "3월",  2023, 3),
    # 4월 has only Sum row (131); skip — no daily data
]


def cv(v):
    """Normalize cell value: None or numeric or string."""
    return v


def num(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return 0


def num_or_null(v):
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return None


def numdec_or_null(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def sql_val(v):
    if v is None:
        return "NULL"
    if isinstance(v, str):
        return "'" + v.replace("'", "''") + "'"
    return str(v)


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.worksheets[0]
    assert ws.title == "INDEX", f"unexpected sheet[0]: {ws.title}"

    rows_sql: list[str] = []
    expected_rows = 0

    for start_row, label, year, month in MONTHS:
        # iterate days 1..31
        for offset in range(0, 31):
            r = start_row + offset
            day_cell = ws.cell(r, 1).value
            dow_cell = ws.cell(r, 2).value
            # stop if we've hit the next month's Sum row or out of days
            if not isinstance(day_cell, int):
                break
            if dow_cell not in DOW_KR_TO_EN:
                break

            try:
                date = dt.date(year, month, day_cell)
            except ValueError:
                # day out of month range
                break

            # Use calendar-derived DOW (sheet's MS labels are inconsistent in Feb).
            calc_dow_kr = ["월", "화", "수", "목", "금", "토", "일"][date.weekday()]
            dkp_dow_en = DOW_KR_TO_EN[calc_dow_kr]
            dkp_dow_kr = calc_dow_kr

            mkt_visitor = num_or_null(ws.cell(r, 3).value)
            mkt_cost = num_or_null(ws.cell(r, 4).value)
            mkt_effect = num_or_null(ws.cell(r, 5).value)
            cs_counseling = num(ws.cell(r, 6).value)
            cs_apply = num(ws.cell(r, 7).value)
            cs_beginning = num(ws.cell(r, 8).value)
            cs_missing = num(ws.cell(r, 9).value)
            cs_trial_class = num(ws.cell(r, 10).value)
            cs_complain = num(ws.cell(r, 11).value)
            ops_new_st = num(ws.cell(r, 12).value)
            ops_out_st = num(ws.cell(r, 13).value)
            ops_count_st = num(ws.cell(r, 14).value)
            ops_new_tc = num(ws.cell(r, 15).value)
            ops_out_tc = num(ws.cell(r, 16).value)
            ops_count_tc = num(ws.cell(r, 17).value)
            cls_map_test = num(ws.cell(r, 18).value)
            cls_tt_class = numdec_or_null(ws.cell(r, 19).value) or 0.0
            cls_student = num(ws.cell(r, 20).value)
            cls_teacher = num(ws.cell(r, 21).value)

            row_sql = (
                "  ("
                f"'{ENT_ID}'::uuid, "
                f"'{date.isoformat()}'::date, "
                f"'{date.strftime('%Y-%m')}', "
                f"{day_cell}, "
                f"'{dkp_dow_en}', "
                f"'{dkp_dow_kr}', "
                f"{sql_val(mkt_visitor)}, "
                f"{sql_val(mkt_cost)}, "
                f"{sql_val(mkt_effect)}, "
                f"{cs_counseling}, {cs_apply}, {cs_beginning}, {cs_missing}, {cs_trial_class}, {cs_complain}, "
                f"{ops_new_st}, {ops_out_st}, {ops_count_st}, "
                f"{ops_new_tc}, {ops_out_tc}, {ops_count_tc}, "
                f"{cls_map_test}, {cls_tt_class:.1f}, {cls_student}, {cls_teacher}, "
                "'FRESH', 'COMPLETE', NOW()"
                ")"
            )
            rows_sql.append(row_sql)
            expected_rows += 1

    header = f"""-- =========================================================================
-- 700-seed-dsh-index-data.sql
--   Source: docs/reference/[TPI] Master.xlsx · sheet[0] INDEX
--   Generated by: scripts/build-dsh-index-seed.py
--   Tenant (ent_id): {ENT_ID} (Trinity demo)
--   Months covered: 2022-12 / 2023-01 / 2023-02 / 2023-03  ({expected_rows} daily rows)
--   Idempotent: ON CONFLICT (ent_id, dkp_date) DO UPDATE
-- =========================================================================

INSERT INTO amb_acm_dsh_daily_kpi (
  ent_id, dkp_date, dkp_year_month, dkp_day_of_month, dkp_day_of_week, dkp_day_of_week_kr,
  dkp_marketing_visitor, dkp_marketing_cost, dkp_marketing_effect,
  dkp_cs_counseling, dkp_cs_apply, dkp_cs_beginning, dkp_cs_missing, dkp_cs_trial_class, dkp_cs_complain,
  dkp_ops_new_st, dkp_ops_out_st, dkp_ops_count_st,
  dkp_ops_new_tc, dkp_ops_out_tc, dkp_ops_count_tc,
  dkp_class_map_test, dkp_class_tt_class, dkp_class_student, dkp_class_teacher,
  dkp_computation_status, dkp_data_completeness, dkp_computed_at
) VALUES
"""

    body = ",\n".join(rows_sql)

    footer = """
ON CONFLICT (ent_id, dkp_date) DO UPDATE SET
  dkp_year_month         = EXCLUDED.dkp_year_month,
  dkp_day_of_month       = EXCLUDED.dkp_day_of_month,
  dkp_day_of_week        = EXCLUDED.dkp_day_of_week,
  dkp_day_of_week_kr     = EXCLUDED.dkp_day_of_week_kr,
  dkp_marketing_visitor  = EXCLUDED.dkp_marketing_visitor,
  dkp_marketing_cost     = EXCLUDED.dkp_marketing_cost,
  dkp_marketing_effect   = EXCLUDED.dkp_marketing_effect,
  dkp_cs_counseling      = EXCLUDED.dkp_cs_counseling,
  dkp_cs_apply           = EXCLUDED.dkp_cs_apply,
  dkp_cs_beginning       = EXCLUDED.dkp_cs_beginning,
  dkp_cs_missing         = EXCLUDED.dkp_cs_missing,
  dkp_cs_trial_class     = EXCLUDED.dkp_cs_trial_class,
  dkp_cs_complain        = EXCLUDED.dkp_cs_complain,
  dkp_ops_new_st         = EXCLUDED.dkp_ops_new_st,
  dkp_ops_out_st         = EXCLUDED.dkp_ops_out_st,
  dkp_ops_count_st       = EXCLUDED.dkp_ops_count_st,
  dkp_ops_new_tc         = EXCLUDED.dkp_ops_new_tc,
  dkp_ops_out_tc         = EXCLUDED.dkp_ops_out_tc,
  dkp_ops_count_tc       = EXCLUDED.dkp_ops_count_tc,
  dkp_class_map_test     = EXCLUDED.dkp_class_map_test,
  dkp_class_tt_class     = EXCLUDED.dkp_class_tt_class,
  dkp_class_student      = EXCLUDED.dkp_class_student,
  dkp_class_teacher      = EXCLUDED.dkp_class_teacher,
  dkp_computation_status = EXCLUDED.dkp_computation_status,
  dkp_data_completeness  = EXCLUDED.dkp_data_completeness,
  dkp_computed_at        = EXCLUDED.dkp_computed_at;
"""

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(header + body + footer, encoding="utf-8")
    print(f"OK: {expected_rows} rows -> {OUT.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
