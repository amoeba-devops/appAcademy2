#!/usr/bin/env python3
"""Generate SCH school master seed SQL from `[TPI] Master.xlsx › 학교입학 정보`.

Usage:
  python3 scripts/build-sch-seed.py > sql/acm/200-seed-sch-schools.sql
"""

import sys
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC_XLSX = ROOT / 'docs' / 'reference' / '[TPI] Master.xlsx'
SHEET = '학교입학 정보'
DEMO_ENT = '00000000-0000-0000-0000-000000000001'


def load_schools():
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb[SHEET]
    schools: dict[str, dict[str, str]] = {}
    section = 1  # 1 = 인가 (Jeju international), 2 = 비인가
    last_region: str | None = None

    for r in range(4, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a == '국내 비인가 국제학교 입학전형':
            section = 2
            continue
        if a == '지역':
            continue
        if isinstance(a, str) and a.startswith('*'):
            continue
        if section == 1:
            if isinstance(a, str) and a.strip():
                name = a.strip()
                schools.setdefault(name, {'region': '제주/국제', 'level': 'FOREIGN'})
        else:
            if isinstance(a, str) and a.strip():
                last_region = a.strip()
            if isinstance(b, str) and b.strip():
                name = b.strip().rstrip('\t').strip()
                if name:
                    schools.setdefault(
                        name, {'region': last_region or '국내', 'level': 'FOREIGN'}
                    )
    return schools


def emit_sql(schools: dict[str, dict[str, str]]):
    out = sys.stdout
    out.write('-- AUTO-GENERATED — Seed SCH school master\n')
    out.write('-- Source: docs/reference/[TPI] Master.xlsx › 학교입학 정보\n')
    out.write(f'-- Rows: {len(schools)}\n')
    out.write('-- Idempotent: skips if any school already exists for the demo ent_id\n\n')
    out.write('BEGIN;\n')
    out.write('DO $$\n')
    out.write('DECLARE v_ent_id UUID;\n')
    out.write('BEGIN\n')
    out.write(f"  v_ent_id := '{DEMO_ENT}'::uuid;\n")
    out.write('  IF EXISTS (SELECT 1 FROM amb_acm_sch_school WHERE ent_id = v_ent_id) THEN\n')
    out.write("    RAISE NOTICE 'SCH seed skipped — schools already exist for ent_id %', v_ent_id;\n")
    out.write('    RETURN;\n')
    out.write('  END IF;\n')
    for name, meta in sorted(schools.items()):
        n = name.replace("'", "''")
        rg = meta['region'].replace("'", "''")
        out.write(
            "  INSERT INTO amb_acm_sch_school (ent_id, name, level, region, is_foreign) "
            f"VALUES (v_ent_id, '{n}', 'FOREIGN', '{rg}', TRUE);\n"
        )
    out.write('END$$;\n')
    out.write('COMMIT;\n')


if __name__ == '__main__':
    emit_sql(load_schools())
