#!/usr/bin/env python3
"""
Trinity Academy — TPI seed v2 generator
Document: STUDENT-IMPORT-TASK-2.0.0

Reads docs/reference/TPI 학생 정보.xlsx and produces:
  /tmp/tpi_seed_v2_review.json   — parsed intermediate (human review)
  sql/seed-tpi-students-v2.sql   — executable seed

Run: python3 scripts/build-seed-v2.py
"""
from __future__ import annotations
import json
import re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "docs/reference/TPI 학생 정보.xlsx"
JSON_OUT = Path("/tmp/tpi_seed_v2_review.json")
SQL_OUT = ROOT / "sql/seed-tpi-students-v2.sql"

WEEKDAY_KO = {'월':'MON','화':'TUE','수':'WED','목':'THU','금':'FRI','토':'SAT','일':'SUN'}


# ---------- normalization helpers ----------

def sql_str(v):
    if v is None:
        return "NULL"
    s = str(v).replace("\\", "\\\\").replace("'", "''")
    return f"'{s}'"


def sql_date(v):
    return "NULL" if v is None else f"DATE '{v}'"


def norm_gender(v):
    if v == '남': return 'M'
    if v == '여': return 'F'
    return None


def parse_date_cell(v):
    """Accepts datetime, numeric (YYYYMMDD), or string variants."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)):
        n = int(v)
        s = str(n)
        if len(s) == 8:  # YYYYMMDD
            try:
                return datetime.strptime(s, '%Y%m%d').strftime('%Y-%m-%d')
            except ValueError:
                return None
        if len(s) == 7:  # e.g. 2017117 — assume single-digit month
            try:
                return datetime.strptime(s, '%Y%m%d' if False else '%Y%m%d').strftime('%Y-%m-%d')
            except ValueError:
                pass
            try:
                y, m, d = s[:4], s[4:5], s[5:]
                return datetime.strptime(f'{y}{m.zfill(2)}{d}', '%Y%m%d').strftime('%Y-%m-%d')
            except ValueError:
                return None
        return None  # year-only or partial
    s = str(v).strip()
    m = re.match(r'^(\d{4})(\d{2})[/-](\d{1,2})$', s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{int(m.group(3)):02d}"
    m = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def parse_phone(v):
    """Returns digits-only string ('01045621469') or None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        digits = str(int(v))
    else:
        digits = re.sub(r'\D', '', str(v))
    if not digits:
        return None
    if len(digits) == 10 and digits.startswith('10'):
        digits = '0' + digits
    if len(digits) in (10, 11) and digits.startswith('0'):
        return digits
    return None


def parse_grade(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return str(int(v))
    return str(v).strip()


def parse_name(raw):
    """
    Returns dict: name, english_name, residence, grade, is_placeholder, raw
    Patterns handled:
      '홍길동(Jamy)'        → name=홍길동, english_name=Jamy
      '고유진/Chloe'         → name=고유진, english_name=Chloe
      '밀라노로 G9, 정가원'  → residence=밀라노, grade=G9, name=정가원
      '듀오링고 G5'          → placeholder
      '밴쿠버 G7'            → placeholder
      'Erica'               → english-only
    """
    s = str(raw).strip()
    out = {"name": s, "english_name": None, "residence": None, "grade": None,
           "is_placeholder": False, "raw": s}

    m = re.match(r'^([^\s]+로)\s*G(\d+)\s*,\s*(.+)$', s)
    if m:
        out["residence"] = m.group(1).rstrip('로')
        out["grade"] = f"G{m.group(2)}"
        out["name"] = m.group(3).strip()
        return out

    m = re.match(r'^(.+?)\s*\(([A-Za-z][^)]*)\)\s*$', s)
    if m:
        out["name"] = m.group(1).strip()
        out["english_name"] = m.group(2).strip()
        return out

    m = re.match(r'^(.+?)\s*/\s*([A-Za-z][A-Za-z ]*)\s*$', s)
    if m:
        out["name"] = m.group(1).strip()
        out["english_name"] = m.group(2).strip()
        return out

    # Placeholder: known location/service names with grade marker but no person name
    if re.match(r'^(듀오링고|밴쿠버|밀라노)\s*G\d+\s*$', s):
        out["is_placeholder"] = True
        return out

    return out


def parse_map_score(v):
    """Returns dict with reading/math/language or None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        n = int(v)
        s = str(n)
        if len(s) == 3:
            return {"reading": n, "math": None, "language": None}
        if len(s) == 6:
            return {"reading": int(s[:3]), "math": int(s[3:]), "language": None}
        return None
    raw = str(v).strip()
    # '234,223(리딩/랭귀지아트)'
    m = re.match(r'^\s*(\d{3})\s*,\s*(\d{3})\s*\(리딩\s*/\s*랭귀지', raw)
    if m:
        return {"reading": int(m.group(1)), "math": None, "language": int(m.group(2))}
    # Fall back: extract any 3-digit numbers
    nums = [int(x) for x in re.findall(r'\b\d{3}\b', raw)]
    if len(nums) >= 2:
        return {"reading": nums[0], "math": nums[1], "language": nums[2] if len(nums)>2 else None}
    if len(nums) == 1:
        return {"reading": nums[0], "math": None, "language": None}
    return None


def parse_schedule(v):
    if not v:
        return None
    m = re.match(r'^\s*([월화수목금토일])\s*(\d{1,2}:\d{2})\s*[-~]\s*(\d{1,2}:\d{2})', str(v).strip())
    if m:
        return [{"weekday": WEEKDAY_KO[m.group(1)], "start": m.group(2), "end": m.group(3)}]
    return [{"raw": str(v).strip()}]


# ---------- sheet 4 parser ----------

def parse_sheet4(wb):
    ws = wb['구 학생 정보']
    rows = []
    last_student_idx = None  # index in `rows` for continuation attachment
    for r in range(3, ws.max_row + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, 16)]
        v_mark, no_val, name, gender, phone, birth, school, grade, residence, \
            map_raw, curriculum, teacher, subject, materials, schedule = cells

        # Detect continuation row: no name, no No., but has teacher or materials
        if not name and not no_val and (teacher or materials):
            if last_student_idx is not None:
                rows[last_student_idx].setdefault("extra_assignments", []).append({
                    "teacher": teacher,
                    "subject": subject,
                    "materials": materials,
                })
            continue

        if not name:
            continue

        parsed_name = parse_name(name)
        map_score = parse_map_score(map_raw)

        row = {
            "src_sheet": "구 학생 정보",
            "src_row": r,
            "no": int(no_val) if isinstance(no_val, (int, float)) else None,
            "active": v_mark == 'v',
            "name": parsed_name["name"],
            "english_name": parsed_name["english_name"],
            "residence": parsed_name["residence"] or (residence if residence else None),
            "is_placeholder": parsed_name["is_placeholder"],
            "gender": norm_gender(gender),
            "phone": parse_phone(phone),
            "phone_raw": str(phone).strip() if phone else None,
            "birth_date": parse_date_cell(birth),
            "birth_raw": str(birth) if birth is not None else None,
            "school": school.strip() if isinstance(school, str) else school,
            "grade": parse_grade(parsed_name["grade"] or grade),
            "map_score": map_score,
            "curriculum": curriculum,
            "teacher": teacher,
            "subject": subject,
            "materials": materials,
            "schedule": schedule,
            "schedule_json": parse_schedule(schedule),
            "extra_assignments": [],
        }
        rows.append(row)
        last_student_idx = len(rows) - 1
    return rows


# ---------- sheet 2 parser ----------

TEMPLATE_MARKERS = [
    '> 변동시', '> 변경시', '> 추가시', '기존 성적 삭제 X',
    '상담내용 (상담월기록)', '상담 예정일 기입', '상담 완료시 실제 상담',
]

def is_template_text(v):
    if not v:
        return False
    s = str(v)
    return any(m in s for m in TEMPLATE_MARKERS)


def parse_sheet2(wb):
    ws = wb['학부모 및 학생 상담']
    rows = []
    for r in range(4, ws.max_row + 1):
        name = ws.cell(row=r, column=3).value
        if not name:
            continue
        cells = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        start_raw, no_val, _, gender, mob, gpa, map_v, ssat, teacher, \
            curriculum, note, goals, satisfaction, last_counsel = cells

        parsed_name = parse_name(name)

        def real(v):
            return None if is_template_text(v) else v

        rows.append({
            "src_sheet": "학부모 및 학생 상담",
            "src_row": r,
            "no": int(no_val) if isinstance(no_val, (int, float)) else None,
            "start_date": parse_date_cell(start_raw),
            "name": parsed_name["name"],
            "english_name": parsed_name["english_name"],
            "gender": norm_gender(gender),
            "mobility": real(mob),
            "gpa": real(gpa),
            "map_note": real(map_v),
            "ssat_isee_note": real(ssat),
            "teacher": teacher,
            "curriculum": real(curriculum),
            "special_note": real(note),
            "goals_note": real(goals),
            "satisfaction_note": real(satisfaction),
            "last_counsel_date": parse_date_cell(last_counsel) if last_counsel and not is_template_text(last_counsel) else None,
        })
    return rows


# ---------- SQL emitter ----------

def emit_sql(sheet4, sheet2):
    lines = []
    P = lines.append

    P("-- =============================================================")
    P("-- Trinity Academy — TPI Student Import v2 (Sheets 2 & 4)")
    P("-- Document: STUDENT-IMPORT-TASK-2.0.0")
    P("-- Generated by: scripts/build-seed-v2.py")
    P("-- Depends: seed-dev.sql, migration-student-import-1.0.0.sql,")
    P("--          seed-tpi-students.sql (v1.0.0)")
    P("-- =============================================================")
    P("")
    P("SET NAMES utf8mb4;")
    P("SET @academy_id = (SELECT acd_id FROM tac_academies WHERE acd_business_registration_no='123-45-67890' LIMIT 1);")
    P("SET @placeholder_parent_id = (SELECT prt_id FROM tac_parents WHERE acd_id=@academy_id AND prt_name='[MIGRATED] Unknown Guardian' LIMIT 1);")
    P("SET @dev_key = SHA2('trinity-dev-key',256);")
    P("")

    # --- Teachers (additional) ---
    P("-- -------------------------------------------------------------")
    P("-- 1. Additional teachers (from Sheet 4)")
    P("-- -------------------------------------------------------------")
    # Collect all unique teacher names from both sheets
    all_teachers = set()
    for r in sheet4:
        if r.get("teacher"): all_teachers.add(str(r["teacher"]).strip())
        for ea in r.get("extra_assignments", []):
            if ea.get("teacher"): all_teachers.add(str(ea["teacher"]).strip())
    for r in sheet2:
        if r.get("teacher"): all_teachers.add(str(r["teacher"]).strip())

    # Known teachers already seeded in v1: 김태윤, 정성경, TBD
    V1_TEACHERS = {'김태윤', '정성경', 'TBD'}
    name_to_ama = {
        '김태윤': 'PENDING-TPI-KTY', '정성경': 'PENDING-TPI-JSK',
        '조혜수': 'PENDING-TPI-CHS', '임승희': 'PENDING-TPI-LSH',
        '손민서': 'PENDING-TPI-SMS', '한승희': 'PENDING-TPI-HSH',
        '김경진': 'PENDING-TPI-KKJ',
    }
    for t in sorted(all_teachers):
        if t in V1_TEACHERS:
            continue
        ama = name_to_ama.get(t, f"PENDING-TPI-{abs(hash(t))%100000:05d}")
        P(f"INSERT IGNORE INTO tac_teachers (acd_id, tch_ama_client_id, tch_employment_type, tch_cached_profile)")
        P(f"VALUES (@academy_id, {sql_str(ama)}, 'FULL_TIME', JSON_OBJECT('name',{sql_str(t)}));")
    P("")

    # --- Parents from phones ---
    P("-- -------------------------------------------------------------")
    P("-- 2. Parents from unique phone numbers (Sheet 4)")
    P("--    Siblings sharing a phone → share a prt_id")
    P("-- -------------------------------------------------------------")
    phone_groups = {}  # phone → first_child_name
    for r in sheet4:
        if r.get("phone") and r["phone"] not in phone_groups:
            phone_groups[r["phone"]] = r["name"]
    for phone, first_child in phone_groups.items():
        pname = f"[IMPORTED] Guardian ({first_child})"
        P(f"INSERT INTO tac_parents (acd_id, prt_name, prt_phone_encrypted, prt_preferred_channel)")
        P(f"SELECT @academy_id, {sql_str(pname)}, AES_ENCRYPT({sql_str(phone)}, @dev_key), 'SMS'")
        P(f" WHERE NOT EXISTS (SELECT 1 FROM tac_parents")
        P(f"   WHERE acd_id=@academy_id AND prt_name={sql_str(pname)});")
    P("")

    # --- Sheet 4 students (dedup by name) ---
    P("-- -------------------------------------------------------------")
    P("-- 3. Sheet 4 — students (dedup by std_name)")
    P("--    Strategy: UPDATE NULL fields on existing; INSERT new.")
    P("-- -------------------------------------------------------------")

    for r in sheet4:
        name = r["name"]
        phone = r["phone"]
        if phone:
            child_name = phone_groups[phone]
            parent_name = f"[IMPORTED] Guardian ({child_name})"
            prt_expr = f"(SELECT prt_id FROM tac_parents WHERE acd_id=@academy_id AND prt_name={sql_str(parent_name)} LIMIT 1)"
        else:
            prt_expr = "@placeholder_parent_id"
        # std_curriculum_text ← 수업교재 (textbook list, our best curriculum signal from Sheet 4)
        # std_note           ← placeholder flag, 수업(subject code), extra assignments
        note_extras = []
        if r["is_placeholder"]:
            note_extras.append("[PLACEHOLDER] name inferred from spreadsheet")
        if r.get("subject"):
            note_extras.append(f"수업: {r['subject']}")
        if r.get("curriculum") and r.get("curriculum") != 'MAP TEST':
            note_extras.append(f"커리큘럼 레이블: {r['curriculum']}")
        if r.get("extra_assignments"):
            for ea in r["extra_assignments"]:
                parts = []
                if ea.get('teacher'): parts.append(f"강사={ea['teacher']}")
                if ea.get('subject'): parts.append(f"과목={ea['subject']}")
                if ea.get('materials'): parts.append(f"교재={ea['materials']}")
                note_extras.append("추가 배정: " + " / ".join(parts))
        note = "\n".join(note_extras) if note_extras else None

        lifecycle = 'ACTIVE' if r["active"] else 'TERMINATED'
        cohort = None
        curriculum_text = r.get("materials") if r.get("materials") else None

        # INSERT if not exists (by name); else UPDATE with COALESCE
        P(f"-- [S4 R{r['src_row']}] {name}" + (f" ({r['english_name']})" if r['english_name'] else ""))
        P("INSERT INTO tac_students")
        P("    (acd_id, prt_id, std_name, std_english_name, std_birth_date, std_gender,")
        P("     std_school, std_grade, std_residence, std_cohort_label,")
        P("     std_curriculum_text, std_note, std_phone_encrypted,")
        P("     std_status, std_lifecycle_status)")
        P("SELECT @academy_id,")
        P(f"       {prt_expr},")
        P(f"       {sql_str(name)},")
        P(f"       {sql_str(r['english_name'])},")
        P(f"       {sql_date(r['birth_date'])},")
        P(f"       {sql_str(r['gender'])},")
        P(f"       {sql_str(r['school'])},")
        P(f"       {sql_str(r['grade'])},")
        P(f"       {sql_str(r['residence'])},")
        P(f"       {sql_str(cohort)},")
        P(f"       {sql_str(curriculum_text)},")
        P(f"       {sql_str(note)},")
        P(f"       " + (f"AES_ENCRYPT({sql_str(r['phone'])}, @dev_key)" if r['phone'] else "NULL") + ",")
        P(f"       'ACTIVE',")
        P(f"       {sql_str(lifecycle)}")
        P(" FROM DUAL")
        P(f" WHERE NOT EXISTS (SELECT 1 FROM tac_students")
        P(f"   WHERE acd_id=@academy_id AND std_name={sql_str(name)});")
        P("")
        # UPDATE path — fill NULLs only (for existing match)
        P("UPDATE tac_students SET")
        P(f"    std_english_name  = COALESCE(std_english_name,  {sql_str(r['english_name'])}),")
        P(f"    std_birth_date    = COALESCE(std_birth_date,    {sql_date(r['birth_date'])}),")
        P(f"    std_gender        = COALESCE(std_gender,        {sql_str(r['gender'])}),")
        P(f"    std_school        = COALESCE(std_school,        {sql_str(r['school'])}),")
        P(f"    std_grade         = COALESCE(std_grade,         {sql_str(r['grade'])}),")
        P(f"    std_residence     = COALESCE(std_residence,     {sql_str(r['residence'])}),")
        P(f"    std_curriculum_text = COALESCE(std_curriculum_text, {sql_str(curriculum_text)}),")
        P(f"    std_note          = COALESCE(std_note,          {sql_str(note)}),")
        P(f"    std_phone_encrypted = COALESCE(std_phone_encrypted, " + (f"AES_ENCRYPT({sql_str(r['phone'])}, @dev_key)" if r['phone'] else "NULL") + ")")
        P(f" WHERE acd_id=@academy_id AND std_name={sql_str(name)};")
        P("")

    # --- Sheet 2 new students (only 2 expected) ---
    existing_names = {r["name"] for r in sheet4}
    # Also v1 students
    V1_NAMES = {'강병찬','강소율','구본의','혜리','이재인','정윤아','정윤지','김민','김아이비',
                '이태오','정하율','김지환','김하음','이채현','황채민',
                '이윤건','이윤후','김라희','박지온','정수인','장연우','장연서','석예준','석유준'}

    P("-- -------------------------------------------------------------")
    P("-- 4. Sheet 2 — new students not present in Sheet 1/4")
    P("-- -------------------------------------------------------------")
    for r in sheet2:
        name = r["name"]
        if name in V1_NAMES or name in existing_names:
            continue
        P(f"-- [S2 R{r['src_row']}] {name} (new)")
        P("INSERT INTO tac_students")
        P("    (acd_id, prt_id, std_name, std_english_name, std_gender,")
        P("     std_status, std_lifecycle_status)")
        P("SELECT @academy_id, @placeholder_parent_id,")
        P(f"       {sql_str(name)}, {sql_str(r['english_name'])}, {sql_str(r['gender'])},")
        P("       'ACTIVE', 'CONSULTING'")
        P(" FROM DUAL")
        P(f" WHERE NOT EXISTS (SELECT 1 FROM tac_students")
        P(f"   WHERE acd_id=@academy_id AND std_name={sql_str(name)});")
        P("")

    # Q-L: Sheet 2 gender overwrites Sheet 1 (Sheet 2 is newer)
    P("-- -------------------------------------------------------------")
    P("-- 5. Q-L: Sheet 2 gender is newest — overwrite existing")
    P("-- -------------------------------------------------------------")
    for r in sheet2:
        if r["gender"] is None:
            continue
        P(f"UPDATE tac_students SET std_gender={sql_str(r['gender'])}")
        P(f" WHERE acd_id=@academy_id AND std_name={sql_str(r['name'])};")
    P("")

    # --- MAP scores ---
    P("-- -------------------------------------------------------------")
    P("-- 6. MAP scores (Sheet 4)")
    P("-- -------------------------------------------------------------")
    for r in sheet4:
        ms = r.get("map_score")
        if not ms:
            continue
        reading = ms['reading'] if ms['reading'] is not None else 'NULL'
        math = ms['math'] if ms['math'] is not None else 'NULL'
        lang = ms['language'] if ms['language'] is not None else 'NULL'
        src_row = r["src_row"]
        note_lit = f"Imported from 구 학생 정보 row {src_row}"
        name = r["name"]
        P("INSERT INTO tac_map_scores (std_id, msc_assessed_at, msc_reading_score, msc_math_score, msc_language_score, msc_source, msc_note)")
        P(f"SELECT s.std_id, CURRENT_DATE, {reading}, {math}, {lang}, 'IMPORT', {sql_str(note_lit)}")
        P(f"  FROM tac_students s WHERE s.acd_id=@academy_id AND s.std_name={sql_str(name)}")
        P(f"  AND NOT EXISTS (SELECT 1 FROM tac_map_scores m WHERE m.std_id=s.std_id AND m.msc_source='IMPORT' AND m.msc_note={sql_str(note_lit)});")
    P("")

    # --- Consultations (Sheet 2) ---
    # Multiple students may share a parent (siblings). tac_consultations
    # lacks a student FK, so we encode the student name as a note prefix
    # '[학생: 이름]' to disambiguate per-student rows sharing a prt_id.
    P("-- -------------------------------------------------------------")
    P("-- 7. Sheet 2 — recurring consultation cycle (per student)")
    P("-- -------------------------------------------------------------")
    for r in sheet2:
        name = r["name"]
        prefix = f"[학생: {name}]"
        note_parts = [f"{prefix} TPI 정기 학부모 상담 (2달 1회) — imported from spreadsheet"]
        if r.get("special_note"):
            note_parts.append(f"특이사항: {r['special_note']}")
        if r.get("curriculum"):
            note_parts.append(f"커리큘럼: {r['curriculum']}")
        if r.get("teacher"):
            note_parts.append(f"담당강사: {r['teacher']}")
        note = "\n".join(note_parts)
        guard = f"{prefix}%"
        P("INSERT INTO tac_consultations (acd_id, prt_id, cst_channel, cst_status, cst_note)")
        P(f"SELECT @academy_id, s.prt_id, 'RECURRING', 'OPEN', {sql_str(note)}")
        P(f"  FROM tac_students s WHERE s.acd_id=@academy_id AND s.std_name={sql_str(name)}")
        P("  AND NOT EXISTS (SELECT 1 FROM tac_consultations c")
        P(f"    WHERE c.acd_id=@academy_id AND c.cst_note LIKE {sql_str(guard)});")
    P("")

    # --- Verification ---
    P("-- -------------------------------------------------------------")
    P("-- 8. Verification")
    P("-- -------------------------------------------------------------")
    P("SELECT '— v2 verification —' AS section;")
    P("SELECT COUNT(*) AS total_students FROM tac_students WHERE acd_id=@academy_id;")
    P("SELECT std_lifecycle_status, COUNT(*) FROM tac_students WHERE acd_id=@academy_id GROUP BY 1;")
    P("SELECT COUNT(*) AS parent_rows FROM tac_parents WHERE acd_id=@academy_id;")
    P("SELECT COUNT(*) AS consultations FROM tac_consultations WHERE acd_id=@academy_id;")
    P("SELECT COUNT(*) AS map_scores FROM tac_map_scores;")
    P("SELECT prt_id, COUNT(*) AS siblings FROM tac_students WHERE acd_id=@academy_id")
    P(" GROUP BY prt_id HAVING COUNT(*)>1 ORDER BY siblings DESC LIMIT 10;")
    P("SELECT '✅ v2 seed complete' AS result;")

    return "\n".join(lines) + "\n"


# ---------- main ----------

def main():
    wb = load_workbook(str(XLSX), data_only=True)
    sheet4 = parse_sheet4(wb)
    sheet2 = parse_sheet2(wb)

    summary = {
        "generated_at": datetime.now().isoformat(timespec='seconds'),
        "counts": {
            "sheet4_students": len(sheet4),
            "sheet4_active_v": sum(1 for r in sheet4 if r["active"]),
            "sheet4_with_phone": sum(1 for r in sheet4 if r["phone"]),
            "sheet4_with_birth": sum(1 for r in sheet4 if r["birth_date"]),
            "sheet4_with_map": sum(1 for r in sheet4 if r["map_score"]),
            "sheet4_continuation_rows": sum(len(r["extra_assignments"]) for r in sheet4),
            "sheet4_placeholders": sum(1 for r in sheet4 if r["is_placeholder"]),
            "sheet2_students": len(sheet2),
        },
        "sheet4_rows": sheet4,
        "sheet2_rows": sheet2,
    }
    JSON_OUT.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[review JSON] {JSON_OUT}")
    for k, v in summary["counts"].items():
        print(f"  {k}: {v}")

    sql = emit_sql(sheet4, sheet2)
    SQL_OUT.write_text(sql, encoding="utf-8")
    print(f"[seed SQL]   {SQL_OUT}  ({len(sql.splitlines())} lines)")


if __name__ == "__main__":
    main()
